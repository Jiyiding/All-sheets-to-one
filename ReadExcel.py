import openpyxl
import xlrd
from configparser import ConfigParser

# load exce file
usercfg = openpyxl.load_workbook('makefiles/userCfg.xlsx')
userSheet = usercfg.get_sheet_names()
userSheets = usercfg.get_sheet_by_name(userSheet[0])
SourceFileName = userSheets.cell(row=2, column=2).value
TargetFileName = userSheets.cell(row=3, column=2).value
StartSheet = userSheets.cell(row=4, column=2).value
EndSheet = userSheets.cell(row=5, column=2).value
ProduceCode = userSheets.cell(row=6, column=2).value
Sheetname = userSheets.cell(row=7, column=2).value
TC_Priority = userSheets.cell(row=8, column=2).value

excel_file = openpyxl.load_workbook(SourceFileName)
# acquire sheets
sheet = excel_file.get_sheet_names()


# 获取某一个sheet中的testCaseName\testPurpose
def requireNmPrId(shets):
    tcNm = shets.cell(row=2, column=5).value
    tcPr = shets.cell(row=5, column=2).value
    tcId = shets.cell(row=2, column=2).value
    return tcNm, tcPr, tcId


# 获取目标testcase首尾sheet_index
for num in sheet:
    if (num == StartSheet):
        starIndex = sheet.index(num)
    if (num == EndSheet):
        endIndex = sheet.index(num)

# sheets = {}

# 获取所有sheet中Index\Description\ExpectResult\Name\Purpose，并全部整合到StepAll中
StepIndex = []
StepDescription = []
StepExpectRt = []
TcNm = []
TcPr = []
TcId = []
StepAll = []

for index in range(starIndex, endIndex + 1):
    # sheet + _index = excel_file.
    sheets = excel_file.get_sheet_by_name(sheet[index])
    tcNm, tcPr, tcId = requireNmPrId(sheets)
    TcNm.append(tcNm)
    TcPr.append(tcPr)
    TcId.append(tcId)
    # 获取某一个sheet中Index\Description\ExpectResult
    stepIndex = []
    stepDescription = []
    stepExpectRt = []
    i_step = 0
    i_row = 19
    j_col = 2
    valueIndex = sheets.cell(row=(i_row + i_step), column=j_col).value
    valueDescription = sheets.cell(row=(i_row + i_step), column=(j_col + 1)).value
    valueExpectRt = sheets.cell(row=(i_row + i_step), column=(j_col + 3)).value

    while (valueIndex == (1 + i_step) or valueDescription != None):
        i_step = i_step + 1
        stepIndex.append(valueIndex)
        stepDescription.append(valueDescription)
        stepExpectRt.append(valueExpectRt)
        valueIndex = sheets.cell(row=(i_row + i_step), column=j_col).value
        valueDescription = sheets.cell(row=(i_row + i_step), column=(j_col + 1)).value
        valueExpectRt = sheets.cell(row=(i_row + i_step), column=(j_col + 3)).value

    StepIndex.append(stepIndex)
    StepDescription.append(stepDescription)
    StepExpectRt.append(stepExpectRt)
StepAll.append(StepIndex)
StepAll.append(StepDescription)
StepAll.append(StepExpectRt)
StepAll.append(TcNm)
StepAll.append(TcPr)
StepAll.append(TcId)
print('Read done')

# 接下来：1，如何创建excel并写入数据；2，哪些数据要写；3，合并单元格；4，makefile文件。

'''
# list 切片，区间前闭后开
sheet1 = sheet[5:11]
print("sheet1 = ", sheet1)

TC_001 = excel_file.get_sheet_by_name('TC_001')
print("TC_001_row1= ", TC_001['C19'].value)
print("TC_001_row1= ", TC_001.cell(row=19, column=3).value)
'''
